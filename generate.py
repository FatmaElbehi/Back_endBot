import openpyxl
from flask import Flask, jsonify, request, send_file, url_for, send_from_directory, make_response
from flask_cors import CORS
import random
from openpyxl import Workbook
from datetime import datetime,timedelta
from werkzeug.utils import secure_filename
import mysql.connector
import os
from Chat import get_response
import xlsxwriter
import string


app = Flask(__name__)
CORS(app)


@app.route('/api/predict', methods=['POST'])
def predict():
    text= request.get_json.get("message")
    response=get_response(text)
    message={"answer": response}
    return jsonify(message)

# configure database connection
config = {
    'user': 'root',
    'password': '',
    'database': 'models_case',
    'port': 3306  # update port number as needed
}

# define database table to store Excel files
directory = "I:\Chat\Back_End_API_PFE\TestGPT_API"
@app.route('/api/save-model', methods=['POST'])
def save_Model():

        new_files = []
        #file_path=''
        for filename in os.listdir(directory):
                if filename.endswith('.xlsx'):
                # check if the file is new
                        file_path = os.path.join(directory, filename)
                        if os.path.isfile(file_path):
                                new_files.append(file_path)


                # open the Excel file and select the worksheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # get the data from the worksheet
        data = []
        for row in ws.iter_rows(values_only=True):
                data.append(tuple(row))
                # connect to database
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        # insert data into database table
        cursor.executemany("INSERT INTO stored_cases (`Actions`, `Data`, `Resultat Attendue`,`Accepted/Rejected`) VALUES (%s, %s, %s,%s)", data)
        conn.commit()
        # close database connection
        cursor.close()
        conn.close()
        # remove the Excel file
        os.remove(file_path)
        return "The file saved with success!!!"
    
    
@app.route('/api/save-to-excel', methods=['POST'])
def save_to_excel():
        data = request.get_json()
        
        if 'obligatoire' in data:
              required_field=data['obligatoire']
        else:
              required_field=None

        if 'min_items' in data:
                min_items_data = int(data['min_items'])
        else:
              min_items_data=0
        if 'max_items' in data:
                max_items_data = int(data['max_items'])
        else:
                min_items_data=0

        if str(required_field) == 'False':
            required_option='Optional'
        else:
            required_option='Required'

        if 'visible' in data:
              visible=data['visible']
        else:
              visible=None

        if str(visible) == 'False':
            visible_option='Invisible'
        else:
            visible_option='Visible'
        
        if 'emptyField' in data :
            empty = data['emptyField']
        else:
            empty = None

        if str(empty) == 'False':
            empty_option='not empty'
        else:
            empty_option='empty'   

        module = data['module']
        component_data = data['composant']
        type_component=data['typeComposant']

        if 'typeData' in data:    
            type_data=data['typeData']
        else:
            type_data=None

        if 'checked' and 'defaut2' and 'numItems' and 'min_items' and 'max_items' and 'pItems' and 'checkItem1' and 'checkItem' in data:    
            check_data=data['checked']
            defaut2_data=data['defaut2'] 
            pitems_data=data['pItems']
            checkItem_data=data['checkItem1']
            numItems_data=int(data['numItems'])
            chItem_data=data['checkItem']
        else:
            check_data=None    
            defaut2_data=None
            numItems_data=None
            checkItem_data=None 
            chItem_data= None
            pitems_data=None

        if str(check_data) == 'False':
            check_option='Unchecked'
        else:
            check_option='Checked'


        if type_component == 'Input' and type_data == 'Double':
            min_value = float(data['valeurEntieremin'])
            max_value = float(data['valeurEntieremax'])
        elif type_component == 'Input' and type_data == 'Integer':
            min_value = int(data['valeurEntieremin'])
            max_value = int(data['valeurEntieremax'])
        

        if 'Datemin' in data:
            minDate = datetime.strptime(data['Datemin'], "%Y-%m-%d")
            maxDate = datetime.strptime(data['Datemax'], "%Y-%m-%d")
        else:
            minDate = None # or some default value
            maxDate = None


        if 'active' and 'defaut1' and 'defaut3' in data:
            active_field=data['active']
            default3_value = data['defaut3']
            defaut1_field=data['defaut1']
            clic_field=data['clic']
        else:
            active_field=None
            default3_value = None
            defaut1_field=None
            clic_field=None

        if 'defaut' in data:
            default_value = data['defaut']
        else:
            default_value = None

        if 'plageCaracteresmin' in data:
            chaineMin_val=int(data['plageCaracteresmin'])
            chaineMax_val=int(data['plageCaracteresmax'])
            lettremin_val=data['lettreMin']
            lettremaj_val=data['lettreMaj']
            chiffres_val=data['chiffres']
            cSpec_val=data['caractereSpec']
            sensible_val=data['caractereSensible']
            espace_val=data['espace']
        else:
            chaineMin_val = None 
            chaineMax_val = None
            lettremin_val = None
            lettremaj_val = None
            chiffres_val = None
            cSpec_val = None
            sensible_val = None
            espace_val = None

        if str(lettremin_val) == 'False':
            lettremin_option='Uppercase'
        else:
            lettremin_option='Lowercase'
            
        if str(lettremaj_val) == 'False':
            lettremaj_option='Lowercase'
        else:
            lettremaj_option='Uppercase'  
                
        if str(chiffres_val) == 'False':
            chiffres_option='No numbers'
        else:
            chiffres_option='Contains Numbers'

        if str(cSpec_val) == 'False':
            spec_option='No special caractere'
        else:
            spec_option='Contains special caractere'

        if str(sensible_val) == 'False':
            caseS_option='No case sensitive'
        else:
            caseS_option='It has a case sensitive' 

        if str(espace_val) == 'False':
            spec_option='No space!'
        else:
            spec_option='Contains space!' 

        if str(active_field) == 'False':
            active_option='Inactive!'
        else:
            active_option='Active!'
        
        if str(clic_field) == 'False':
            cli_option='Not cliquable!'
        else:
            cli_option='Cliquable!'


        workbook = xlsxwriter.Workbook(f'{module}.xlsx')
        worksheet_names = set(workbook.sheetnames)

        # group component data by component name
        component_data_by_name = {}

        for component in component_data:
            if "component_name" in component:
                component_name = component["component_name"]
            else:
                component_name = data['composant']

            if component_name not in component_data_by_name:
                component_data_by_name[component_name] = []
            component_data_by_name[component_name].append(component)
        for component_name, component_data in component_data_by_name.items():
            test_cases = []
            if type_component == 'Input' and type_data == 'Integer' :
                for values in [
                    [min_value],
                    [max_value],
                    [min_value - 1],
                    [max_value - 1],
                    [random.randint(min_value, max_value)],
                    [max_value + 1],
                    [min_value + 1]
                ]:
                
                    for i in range(8):
                        out_of_range = False
                        value_description = ' '
                        for value in values:
                            if value < min_value or value > max_value:
                                out_of_range = True
                                break
                        if out_of_range:
                            action = "Enter a value less than "+ str(min_value)+" or greater than "+ str(max_value)+" in the "+component_name+" field"
                            expected_result = f"The entered value is {values} will be rejected because it should be between ["+str(min_value) +","+str(max_value)+"]"
                        else:
                            if len(values) == 1:
                                value_description += f"Value equals {values[0]}"
                            elif len(values) == 2:
                                value_description += f"Value is between {values[0]} and {values[1]}"
                            else:
                                value_description += f"Value is {values}"
                            
                            required=random.choice([True, False])
                            visible_field=random.choice([True, False])
                            empty_field=random.choice([True, False])
                            if default_value:
                                default_str = f"{default_value}"
                            else:
                                default_str = "0"
                            
                            if required and visible_field and empty_field:
                                    value_description += ' - Required field with a visible content and by default: it has an empty value '
                            elif not required and not visible_field and not empty_field :
                                    value_description +=f' - Optional field with content invisible! and by default: it contains this value: {default_str}'
                            elif required and not visible_field and not empty_field:
                                    value_description +=f' - Required field with content invisible! and by default: it contains this value: {default_str}'
                            elif not required and visible_field and empty_field:
                                    value_description += ' - Optional field with a visible content and by default: it has an empty value'
                            elif required and not visible_field and empty_field:
                                    value_description += ' - Required field with a invisible content and by default: it has an empty value'
                            elif required and visible_field and not empty_field:
                                    value_description += f' - Required field with a visible content and by default: it contains this value: {default_str}'
                            elif not required and visible_field and not empty_field:
                                    value_description += f' - Optional field with a visible content and by default: it contains this value: {default_str}'
                            elif not required and not visible_field and empty_field:
                                    value_description += ' - Optional field with invisible content and by default: it has an empty value'
                            
                                # Create test cases for when the field is required and when it is not required
                            if required != bool(required_field) and visible_field != bool(visible) and empty_field != bool(empty)  :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                                # Create test cases for when the field is visible and when it is not visible
                            elif required != bool(required_field) and visible_field != bool(visible) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field '
                            elif required != bool(required_field) and empty_field != bool(empty)  :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field and should be a '+ str(empty_option) +' field'
                            elif visible_field != bool(visible) and empty_field != bool(empty)  :
                                    value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                            
                            elif visible_field != bool(visible) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field'
                            elif required != bool(required_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field'
                                
                            elif empty_field != bool(empty) :
                                    value_description += ' - Rejected Case: because it should be '+ str(empty_option) +' field'
                            
                            elif not empty:
                                value_description = f"{value_description}, and by default: it contains this value: {default_value}"
                            
                            action = f"Add the value {values} in the {component_name} input field"
                            expected_result = value_description

                        data_str = f"Module: {module}"
                        if component_name is not None:
                            data_str += f", Component Name: {component_name}"
                            data_str += f", Value: {values}"
                        else:
                            data_str += f", Value: {values}"

                        if "Rejected" in expected_result:
                                rejected_accepted_field = "Rejected"
                        else:
                                rejected_accepted_field = "Accepted"   

                        test_case = {"Action": action, "Data": data_str, "Expected Result": expected_result,"Rejected/Accepted":rejected_accepted_field}
                        test_cases.append(test_case)
            if type_component == 'Input' and type_data == 'Double' :
                    for values in [
                        [round((min_value),2)],
                        [round((max_value),2)],
                        [round((min_value - 0.5),2)],
                        [round((max_value - 0.5),2)],
                        [round(random.uniform(min_value, max_value),2)],
                        [max_value + 0.5],
                        [min_value + 0.5]
                    ]:
                        for i in range(8):
                            out_of_range = False
                            value_description = ' '
                            for value in values:
                                if value < min_value or value > max_value:
                                    out_of_range = True
                                    break
                            if out_of_range:
                                action = "Enter a value less than "+ str(min_value)+" or greater than "+ str(max_value)+" in the "+component_name+" field"
                                expected_result = f"The entered value is {values} will be rejected because it should be between ["+str(min_value) +","+str(max_value)+"]"
                            else:
                                if len(values) == 1:
                                    value_description += f"Value equals {values[0]}"
                                elif len(values) == 2:
                                    value_description += f"Value is between {values[0]} and {values[1]}"
                                else:
                                    value_description += f"Value is {values}"
                                
                                required=random.choice([True, False])
                                visible_field=random.choice([True, False])
                                empty_field=random.choice([True, False])
                                if default_value:
                                    default_str = f"{default_value}"
                                else:
                                    default_str = "0"
                                
                                if required and visible_field and empty_field:
                                        value_description += ' - Required field with a visible content and by default: it has an empty value '
                                elif not required and not visible_field and not empty_field :
                                        value_description +=f' - Optional field with content invisible! and by default: it contains this value: {default_str}'
                                elif required and not visible_field and not empty_field:
                                        value_description +=f' - Required field with content invisible! and by default: it contains this value: {default_str}'
                                elif not required and visible_field and empty_field:
                                        value_description += ' - Optional field with a visible content and by default: it has an empty value'
                                elif required and not visible_field and empty_field:
                                        value_description += ' - Required field with a invisible content and by default: it has an empty value'
                                elif required and visible_field and not empty_field:
                                        value_description += f' - Required field with a visible content and by default: it contains this value: {default_str}'
                                elif not required and visible_field and not empty_field:
                                        value_description += f' - Optional field with a visible content and by default: it contains this value: {default_str}'
                                elif not required and not visible_field and empty_field:
                                        value_description += ' - Optional field with invisible content and by default: it has an empty value'
                                
                                    # Create test cases for when the field is required and when it is not required
                                if required != bool(required_field) and visible_field != bool(visible) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                                    # Create test cases for when the field is visible and when it is not visible
                                elif required != bool(required_field) and visible_field != bool(visible) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field '
                                elif required != bool(required_field) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field and should be a '+ str(empty_option) +' field'
                                elif visible_field != bool(visible) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                                
                                elif visible_field != bool(visible) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field'
                                elif required != bool(required_field) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field'
                                    
                                elif empty_field != bool(empty) :
                                        value_description += ' - Rejected Case: because it should be '+ str(empty_option) +' field'
                                
                                elif not empty:
                                    value_description = f"{value_description}, and by default: it contains this value: {default_value}"
                                
                                action = f"Add the value {values} in the {component_name} input field"
                                expected_result = value_description

                            data_str = f"Module: {module}"
                            if component_name is not None:
                                data_str += f", Component Name: {component_name}"
                                data_str += f", Value: {values}"
                            else:
                                data_str += f", Value: {values}"

                            if "Rejected" in expected_result:
                                rejected_accepted_field = "Rejected"
                            else:
                                rejected_accepted_field = "Accepted"   

                            test_case = {"Action": action, "Data": data_str, "Expected Result": expected_result,"Rejected/Accepted":rejected_accepted_field}
                            test_cases.append(test_case)                           

            if type_component == 'Input' and type_data == 'Date' :
                    for values in [ 
                        [minDate], # format the dates according to the selected format
                        [maxDate],
                        [(minDate - timedelta(days=1))],
                        [(maxDate - timedelta(days=1))],
                        [(minDate + timedelta(seconds=random.randrange(int((maxDate - minDate).total_seconds()))))],
                        [(maxDate + timedelta(days=1))],
                        [(minDate + timedelta(days=1))]
                        ]:
                        for i in range(8):
                            out_of_range = False
                            value_description = ' '
                            for value in values:
                                if value < minDate or value > maxDate:
                                    out_of_range = True
                                    break
                            if out_of_range:
                                action = "Enter a value less than "+ str(minDate)+" or greater than "+ str(maxDate)+" in the "+component_name+" field"
                                expected_result = f"The entered value is {values} will be rejected because it should be between ["+str(minDate) +","+str(maxDate)+"]"
                            else:
                                if len(values) == 1:
                                    value_description += f"Value equals {values[0]}"
                                elif len(values) == 2:
                                    value_description += f"Value is between {values[0]} and {values[1]}"
                                else:
                                    value_description += f"Value is {values}"
                                
                                required=random.choice([True, False])
                                visible_field=random.choice([True, False])
                                empty_field=random.choice([True, False])

                                if default_value:
                                    default_str = f"{default_value}"
                                else:
                                    default_str = "0"
                                
                                if required and visible_field and empty_field:
                                        value_description += ' - Required field with a visible content and by default: it has an empty value '
                                elif not required and not visible_field and not empty_field :
                                        value_description +=f' - Optional field with content invisible! and by default: it contains this value: {default_str}'
                                elif required and not visible_field and not empty_field:
                                        value_description +=f' - Required field with content invisible! and by default: it contains this value: {default_str}'
                                elif not required and visible_field and empty_field:
                                        value_description += ' - Optional field with a visible content and by default: it has an empty value'
                                elif required and not visible_field and empty_field:
                                        value_description += ' - Required field with a invisible content and by default: it has an empty value'
                                elif required and visible_field and not empty_field:
                                        value_description += f' - Required field with a visible content and by default: it contains this value: {default_str}'
                                elif not required and visible_field and not empty_field:
                                        value_description += f' - Optional field with a visible content and by default: it contains this value: {default_str}'
                                elif not required and not visible_field and empty_field:
                                        value_description += ' - Optional field with invisible content and by default: it has an empty value'
                                
                                    # Create test cases for when the field is required and when it is not required
                                if required != bool(required_field) and visible_field != bool(visible) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                                    # Create test cases for when the field is visible and when it is not visible
                                elif required != bool(required_field) and visible_field != bool(visible) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field '
                                elif required != bool(required_field) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field and should be a '+ str(empty_option) +' field'
                                elif visible_field != bool(visible) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                                
                                elif visible_field != bool(visible) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field'
                                elif required != bool(required_field) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field'
                                    
                                elif empty_field != bool(empty) :
                                        value_description += ' - Rejected Case: because it should be '+ str(empty_option) +' field'
                                
                                elif not empty:
                                    value_description = f"{value_description}, and by default: it contains this value: {default_value}"
                                
                                action = f"Add the value {values} in the {component_name} input field"
                                expected_result = value_description

                            data_str = f"Module: {module}"
                            if component_name is not None:
                                data_str += f", Component Name: {component_name}"
                                data_str += f", Value: {values}"
                            else:
                                data_str += f", Value: {values}"

                            if "Rejected" in expected_result:
                                        rejected_accepted_field = "Rejected"
                            else:
                                        rejected_accepted_field = "Accepted"   

                            test_case = {"Action": action, "Data": data_str, "Expected Result": expected_result,"Rejected/Accepted":rejected_accepted_field}
                            test_cases.append(test_case)                
            
            if type_component == 'Input' and type_data == 'String' :
                    for values in [
                        [chaineMin_val],
                        [chaineMax_val],
                        [chaineMin_val - 1],
                        [chaineMax_val - 1],
                        [random.randint(chaineMin_val, chaineMax_val)],
                        [chaineMax_val + 1],
                        [chaineMin_val + 1]
                    ]:
                        for i in range(8):
                            out_of_range = False
                            value_description = ' '
                            length = random.randint(chaineMin_val, chaineMax_val)
                            random_string = ''.join(random.choices(string.ascii_letters + string.digits, k=length))
                            for value in values:
                                if len(random_string) < chaineMin_val or len(random_string) > chaineMax_val:
                                    out_of_range = True
                                    break
                                if out_of_range:
                                    action =f"Enter a string of length between {chaineMin_val} and {chaineMax_val} in the {component_name} field"
                                    expected_result = f"The entered value is {random_string} will be rejected because its length is not between {chaineMin_val} and {chaineMax_val}"
                                else:
                                    if len(values) == 1:
                                        value_description += f"String equals [{random_string}] of length between {chaineMin_val} and {chaineMax_val}"
                                    else:
                                        value_description += f"Value is {values}"
                                    
                                required=random.choice([True, False])
                                visible_field=random.choice([True, False])
                                empty_field=random.choice([True, False])
                                lowercase_field=random.choice([True, False])
                                upper_field=random.choice([True, False])
                                nb_field=random.choice([True, False])
                                sCarac_field=random.choice([True, False])
                                caseSens_field=random.choice([True, False])
                                space_field=random.choice([True, False])

                                if default_value:
                                    default_str = f"{default_value}"
                                else:
                                    default_str = "0"
                                
                                if required and visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description += ' - Required field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters, it contains numbers.It accepts space with case sensitive'
                                elif not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, it doesn't contain lower no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Required field with content invisible! and by default: it contains this word: {default_str}, it doesn't contain lower no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif not required and visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description += ' - Optional field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters, it contains numbers.It accepts space with case sensitive'
                                elif  required and visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Required field with a visible content and by default: it contains this word: {default_str}, it doesn't contain lower no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  required and visible_field and empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Required field with a visible content and by default: it has an empty string, it doesn't contain lower no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  required and visible_field and empty_field and lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Required field with a visible content and by default: it has an empty string, it contains lower but no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  required and visible_field and empty_field and lowercase_field and upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Required field with a visible content and by default: it has an empty string, it contains lower and upper caracteres but no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  required and visible_field and empty_field and lowercase_field and upper_field and nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Required field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and numbers but no special caracters.It doesn't accept space and no case sensitive"
                                elif  required and visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Required field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters ,numbers.It doesn't accept space and no case sensitive."
                                elif  required and visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Required field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters,numbers and case sensitive.It doesn't accept space."
                                elif  not required and visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it contains this word: {default_str}, it doesn't contain lower no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and visible_field and empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it doesn't contain lower no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and visible_field and empty_field and lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it contains lower but no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and visible_field and empty_field and lowercase_field and upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it contains lower and upper caracteres but no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and visible_field and empty_field and lowercase_field and upper_field and nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and numbers but no special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters ,numbers.It doesn't accept space and no case sensitive."
                                elif  not required and visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters,numbers and case sensitive.It doesn't accept space."
                                elif  not required and visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters,numbers and case sensitive.It accepts space."
                                
                                elif  not required and not visible_field and empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it has an empty string, it doesn't contain lower no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and empty_field and lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it has an empty string, it contains lower but no upper caracteres , no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and empty_field and lowercase_field and upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it has an empty string, it contains lower and upper caracteres , but no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and empty_field and lowercase_field and upper_field and nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it has an empty string, it contains lower and upper caracteres no numbers, but no special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it has an empty string, it contains lower, upper caracteres, numbers and special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it has an empty string, it contains lower, upper caracteres and special caracters ,numbers, special caracters and case sensitive.It doesn't accept space."
                                elif  not required and not visible_field and empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description +=f" - Optional field with a visible content and by default: it has an empty string, it contains lower, upper caracteres and special caracters,numbers and case sensitive.It doesn't accept spaces."
                            
                                elif  not required and not visible_field and not empty_field and lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, it contains lower caracteres but no upper and no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, it contains lower but and upper caracteres, no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, it contains lower and upper caracteres and numbers, but no special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and nb_field and sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, it contains lower and upper caracteres and numbers also special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, it contains lower, upper caracteres, numbers and special caracters also case sensitive.It doesn't accept spaces"
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, it contains lower, upper caracteres and special caracters ,numbers, special caracters and case sensitive.It accepts spaces."
                                
                                elif  not required and not visible_field and not empty_field and not lowercase_field and upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower caracteres but it contains upper and no special caracters and no numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and not lowercase_field and upper_field and nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower caracteres but it contains upper caracteres and numbers, no special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and nb_field and sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower caracteres but it contains upper caracteres, numbers and special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower caracteres but it contains upper caracteres, numbers also special caracters and case sensitive.It doesn't accept spaces."
                                elif  not required and not visible_field and not empty_field and lowercase_field and upper_field and nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower caracteres but it contains upper caracteres, numbers and special caracters and case sensitive.It accepts spaces"
                            
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no special caracters.It contains numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and nb_field and sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres.It contains special caracters and numbers.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and nb_field and sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres. It contains numbers, special caracters and case sensitive.It doesn't accept spaces"
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and upper caracteres. It contains numbers also special caracters and case sensitive.It accepts spaces."
                            
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no numbers.It contains special caracters.It doesn't accept space and no case sensitive"
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no numbers.It contains special caracters and case sensitive.It doesn't accept spaces."
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and sCarac_field and caseSens_field and space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no numbers. It contains special caracters and case sensitive. It accepts spaces"
                                
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no numbers and no special caracters.It has case sensitive. It doesn't accept spaces."
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and caseSens_field and space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no numbers and no special caracters.It has case sensitive.It accepts spaces."
           
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and not space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no numbers, no special caracters and no case sensitive. It doesn't accept spaces."
                                elif  not required and not visible_field and not empty_field and not lowercase_field and not upper_field and not nb_field and not sCarac_field and not caseSens_field and space_field :
                                        value_description +=f" - Optional field with content invisible! and by default: it contains this word: {default_str}, no lower and no upper caracteres, no numbers no special caracters, and no case sensitive.It accepts spaces."                            
                                    
                                    # Create test cases for when the field is required and when it is not required
                                if required != bool(required_field) and visible_field != bool(visible) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                                    # Create test cases for when the field is visible and when it is not visible
                                elif required != bool(required_field) and visible_field != bool(visible) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(visible_option) +' field '
                                elif required != bool(required_field) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field and should be a '+ str(empty_option) +' field'
                                elif visible_field != bool(visible) and empty_field != bool(empty)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field and should be a '+ str(empty_option) +' field'
                                
                                elif visible_field != bool(visible) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(visible_option) +' field'
                                elif required != bool(required_field) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field'
                                    
                                elif empty_field != bool(empty) :
                                        value_description += ' - Rejected Case: because it should be '+ str(empty_option) +' field'
                                
                                elif lowercase_field != bool(lettremin_val) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(lettremin_option) +' field'

                                elif upper_field != bool(lettremaj_val) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(lettremaj_option) +' field'

                                elif nb_field != bool(chiffres_val) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(chiffres_option) +' field'
                                elif sCarac_field != bool(cSpec_val) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(spec_option) +' field'
                                elif caseSens_field != bool(sensible_val) : 
                                        value_description += ' - Rejected Case: because it should be a '+ str(caseS_option) +' field'
                                elif space_field != bool(espace_val) : 
                                        value_description += ' - Rejected Case: because it should be a '+ str(spec_option) +' field'

                                elif not empty:
                                    value_description = f"{value_description}, and by default: it contains this word: {default_value}"
                                
                                action = f"Add the value [{random_string}] in the {component_name} input field"
                                expected_result = value_description

                            data_str = f"Module: {module}"
                            if component_name is not None:
                                data_str += f", Component Name: {component_name}"
                                data_str += f", Value: {random_string}"
                            else:
                                data_str += f", Value: {random_string}"

                            if "Rejected" in expected_result:
                                rejected_accepted_field = "Rejected"
                            else:
                                rejected_accepted_field = "Accepted"   

                            test_case = {"Action": action, "Data": data_str, "Expected Result": expected_result,"Rejected/Accepted":rejected_accepted_field}
                            test_cases.append(test_case)
            if type_component == 'Button':
                for i in range(8):
                            action = "Clic on the button named by: "+ str(component_name)+"."
                            expected_result = "The system will be redirect to the desired page."
                            value_description = str(component_name) +' '
                            required=random.choice([True, False])
                            active=random.choice([True, False])
                            clic=random.choice([True, False])

                            if default3_value:
                                default_str = f"{default3_value}"
                            else:
                                default_str = "0"

                            if defaut1_field:
                                    default_str1 = f"{defaut1_field}"
                            else:
                                    default_str1 = "The previous inputs will be filled!"

                            if not required and not active and not clic:
                                    print(f"default_str = {default_str}")
                                    value_description += f' - Optional button, with inactive state!! ==> when: {default_str} and not clickable!! when ==> {default_str1}'
                            elif not required and not active and not clic :
                                    value_description +=f' - Optional button, with inactive state!! ==> when: {default_str} and not clickable!! when ==> {default_str1}'
                            elif not required and active and clic:
                                    value_description +=f' - Optional,active when ==> {default_str} and clickable button, when ==> {default_str1}'
                            elif not required and not active and clic:
                                    value_description +=f' - Optional and clickable button when ==> {default_str1}, with inactive state!! ==> when: {default_str}.'
                            elif not required and active and not clic:
                                    value_description +=f' - Optional and active Button when {default_str} and not clickable  !! when ==> {default_str1}'
                                
                            elif required and not active and not clic:
                                    value_description += f' - Required Button with inactive state!! ==> when: {default_str} and not cliquable!! when ==> {default_str1}'
                            elif required and active and not clic:
                                    value_description += f' - Required and active Button when {default_str} and not clickable !! when ==> {default_str1}'
                            elif required and not active and clic:
                                    value_description +=f' - Required and clickable Button when {default_str1} with inactive state!! ==> when: {default_str}'
                
                                    # Create test cases for when the field is required and when it is not required
                                                                # Create test cases for when the field is visible and when it is not visible
                            if required!= bool(required_field) and active != bool(active_field) and  clic != bool(clic_field):
                                    value_description += ' - Accepted Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
                            elif required== bool(required_field) and active == bool(active_field) and  clic == bool(clic_field):
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
                            elif required == bool(required_field) and active != bool(active_field) and clic !=bool(clic_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
                            elif required != bool(required_field) and active == bool(active_field) and  clic ==bool(clic_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
   
                            elif required == bool(required_field) and active == bool(active_field) and clic !=bool(clic_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
                            elif required != bool(required_field) and active != bool(active_field) and clic ==bool(clic_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
   
                            elif required == bool(required_field) and active != bool(active_field) and clic ==bool(clic_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
   
                            elif required != bool(required_field) and active == bool(active_field) and clic !=bool(clic_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button also it should be a '+ str(active_option) +' button and should be a '+ str(cli_option) +' button'
   
                                 
                            if clic != bool(clic_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(cli_option) +' button.'
                            elif required != bool(required_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' button.'
                                    
                            elif active != bool(active_field) :
                                    value_description += ' - Rejected Case: because it should be '+ str(active_option) +' button.'
                                
                            elif not active:
                                    value_description = f"{value_description}, and by default: it contains this value: {default_str}"
                                
                            elif not clic:
                                    value_description = f"{value_description}, and by default: it contains this value: {default_str1}"
                                
                            expected_result = value_description

                            data_str = f"Module: {module}"
                            if component_name is not None:
                                data_str += f", Component Name: {component_name}"

                            if "Rejected" in expected_result:
                                rejected_accepted_field = "Rejected"
                            else:
                                rejected_accepted_field = "Accepted"   

                            test_case = {"Action": action, "Data": data_str, "Expected Result": expected_result,"Rejected/Accepted":rejected_accepted_field}
                            test_cases.append(test_case)
            if type_component == 'Checkbox' :
                    for values in [
                        [min_items_data],
                        [max_items_data],
                        [min_items_data - 1],
                        [max_items_data - 1],
                        [random.randint(min_items_data, max_items_data)],
                        [max_items_data + 1],
                        [min_items_data + 1]
                     ]:
                        for i in range(8):
                            out_of_range = False
                            value_description = ' '
                            for value in values:
                                if value < min_items_data or value > max_items_data:
                                   out_of_range = True
                                   break
                                if out_of_range:
                                   action = "Check number of options less than "+ str(min_items_data)+" or greater than "+ str(max_items_data)+" for the "+component_name+" checkbox"
                                   expected_result = f"The nbr of items checked is {values} will be rejected because it should be between ["+str(min_items_data) +","+str(max_items_data)+"]"
                                else:
                                     if len(values) == 1:
                                            value_description += f"The nbr of items checked equals {values[0]}, which are [{checkItem_data}]"
                                     elif len(values) == 2:
                                            value_description += f"The nbr of items checked is between {values[0]} and {values[1]},which are [{checkItem_data}]"
                                     else:
                                            value_description += f"The nbr of items checked is {values},which are [{checkItem_data}]"
                                
                                required=random.choice([True, False])
                                active=random.choice([True, False])
                                checked=random.choice([True, False])

                                if defaut2_data:
                                        default2_str = f"{defaut2_data}"
                                else:
                                        default2_str = "0"

                                if default3_value:
                                        default3_str = f"{default3_value}"
                                else:
                                        default3_str = "0"
                                
                                if required and active and checked:
                                        value_description +=f' - Required , active content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} as Number of items.'
                                elif not required and not active and not checked:
                                        value_description +=f' - Optional , inactive content when {default3_str} and by default not checked. It contains {numItems_data} as Number of items.'
                                elif required and not active and not checked:
                                        value_description +=f' - Required , inactive content when {default3_str} and by default not checked. It contains {numItems_data} as Number of items.'
                                elif not required and active and checked:
                                        value_description +=f' - Optional , active content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} as Number of items.'

                                elif required and not active and checked:
                                        value_description +=f' - Required , inactive content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} as Number of items.'
                                elif required and active and not checked:
                                        value_description +=f' - Required , active content when {default3_str} and by default not checked. It contains {numItems_data} as Number of items.'
                                elif not required and active and not checked:
                                        value_description +=f' - Optional , active content when {default3_str} and by default not checked. It contains {numItems_data} as Number of items.'
                                elif not required and not active and checked:
                                        value_description +=f' - Optional , inactive content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} as Number of items.'
                                
                                
                                if required != bool(required_field) and active != bool(active_field) and checked != bool(check_data)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(active_option) +' field and should be a '+ str(check_option) +' field'
                                elif required != bool(required_field) and active != bool(active_field) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(active_option) +' field '
                                elif required != bool(required_field) and checked != bool(check_data)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field and should be a '+ str(check_option) +' field'
                                elif checked != bool(check_data) and active != bool(active_field)  :
                                        value_description += ' - Rejected Case: because it should be a '+ str(check_option) +' field and should be a '+ str(active_option) +' field'
                                
                                elif active != bool(active_field) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(active_option) +' field'
                                elif required != bool(required_field) :
                                        value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field'
                                        
                                elif checked != bool(check_data) :
                                        value_description += ' - Rejected Case: because it should be '+ str(check_option) +' field'
                                                                
                                action = f"Check {numItems_data} as number of options in the {component_name} checkbox field which are [{checkItem_data}]"
                                expected_result = value_description

                            data_str = f"Module: {module}"
                            if component_name is not None:
                                data_str += f", Component Name: {component_name}"
                                data_str += f", Value: [{checkItem_data}]"
                                data_str += f", Number of items: [{numItems_data}]" 
                            else:
                                data_str += f", Value: [{checkItem_data}]"

                            if "Rejected" in expected_result:
                                        rejected_accepted_field = "Rejected"
                            else:
                                        rejected_accepted_field = "Accepted"   

                            test_case = {"Action": action, "Data": data_str, "Expected Result": expected_result,"Rejected/Accepted":rejected_accepted_field}
                            test_cases.append(test_case)

        if type_component == 'Radio' :
                    for i in range(10):
                        value_description = ' '
                        expected_result=' '
                        action =f"Tick the item named '{chItem_data}' in this {component_name} compoment where the number of items included here, is: [{numItems_data}]"
                        required=random.choice([True, False])
                        active=random.choice([True, False])
                        checked=random.choice([True, False])

                        values_list = pitems_data.split(",")
                        

                        if defaut2_data:
                                default2_str = f"{defaut2_data}"
                        else:
                                default2_str = random.choice(values_list)

                        if default3_value:
                                default3_str = f"{default3_value}"
                        else:
                                default3_str = "The previous inputs are filled!"

                                                                                       
                        if chItem_data:
                                expected_result = f"Just the item named: {chItem_data} is ticked! while the number of items is equal to [{numItems_data}]"
                        
                        if chItem_data == default2_str and chItem_data!=" ":
                                expected_result = f"The item named: {chItem_data} is ticked and it was the same one set by default!! while the number of items is equal to [{numItems_data}]"
                        
                        elif chItem_data != default2_str :
                                expected_result = f"The item named: {chItem_data} is ticked and it was different than the one set by default!! while the number of items is equal to [{numItems_data}]"
 
                        else :
                                expected_result = f"Rejected --> No item is ticked!!!!"
                        
                        if required and active and checked:
                                    value_description =f'{expected_result} --> Required , active content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} items.'
                        elif not required and not active and not checked:
                                    value_description =f'{expected_result} --> Optional , inactive content when {default3_str} and by default not checked. It contains {numItems_data} items.'
                        elif required and not active and not checked:
                                    value_description =f'{expected_result} --> Required , inactive content when {default3_str} and by default not checked. It contains {numItems_data} items.'
                        elif not required and active and checked:
                                    value_description =f'{expected_result} --> Optional , active content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} items.'

                        elif required and not active and checked:
                                    value_description =f'{expected_result} --> Required , inactive content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} items.'
                        elif required and active and not checked:
                                    value_description =f'{expected_result} --> Required , active content when {default3_str} and by default not checked. It contains {numItems_data} items.'
                        elif not required and active and not checked:
                                    value_description =f'{expected_result} --> Optional , active content when {default3_str} and by default not checked. It contains {numItems_data} items.'
                        elif not required and not active and checked:
                                    value_description =f'{expected_result} --> Optional , inactive content when {default3_str} and by default check this item:{default2_str}. It contains {numItems_data} items.'
                          
                            
                        if required != bool(required_field) and active != bool(active_field) and checked != bool(check_data)  :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(active_option) +' field and should be a '+ str(check_option) +' field'
                        elif required != bool(required_field) and active != bool(active_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field also it should be a '+ str(active_option) +' field '
                        elif required != bool(required_field) and checked != bool(check_data)  :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field and should be a '+ str(check_option) +' field'
                        elif checked != bool(check_data) and active != bool(active_field)  :
                                    value_description += ' - Rejected Case: because it should be a '+ str(check_option) +' field and should be a '+ str(active_option) +' field'
                            
                        elif active != bool(active_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(active_option) +' field'
                        elif required != bool(required_field) :
                                    value_description += ' - Rejected Case: because it should be a '+ str(required_option) +' field'
                                
                        elif checked != bool(check_data) :
                                    value_description += ' - Rejected Case: because it should be '+ str(check_option) +' field'
                                                        
                        expected_result += value_description

                        data_str = f"Module: {module}"
                        if component_name is not None:
                            data_str += f", Component Name: {component_name}"
                            data_str += f", Value: [{chItem_data}]"
                            data_str += f", Number of items: [{numItems_data}]"
                        else:
                            data_str += f", Value: [{chItem_data}]"

                        if "Rejected" in expected_result:
                                        rejected_accepted_field = "Rejected"
                        else:
                                        rejected_accepted_field = "Accepted"   

                        test_case = {"Action": action, "Data": data_str, "Expected Result": expected_result,"Rejected/Accepted":rejected_accepted_field}
                        test_cases.append(test_case)

                # write test cases to worksheet
        if not component_name:
                worksheet_name = "Untitled"
        else:
                worksheet_name = component_name
                count = 1
                while worksheet_name.lower() in [name.lower() for name in worksheet_names]:
                    worksheet_name = f"{component_name}_{count}"
                    count += 1
                worksheet = workbook.add_worksheet(worksheet_name)
                header_format = workbook.add_format({'bold': True, 'bg_color': 'yellow'})
                row = 0
                worksheet.write(row, 0, 'Action', header_format)
                worksheet.write(row, 1, 'Data', header_format)
                worksheet.write(row, 2, 'Expected Result', header_format)
                worksheet.write(row, 3, 'Rejected/Accepted', header_format)
                row += 1
                for test_case in test_cases:
                    worksheet.write(row, 0, test_case['Action'])
                    worksheet.write(row, 1, test_case['Data'])
                    worksheet.write(row, 2, test_case['Expected Result'])
                    worksheet.write(row, 3, test_case['Rejected/Accepted'])
                    row += 1
                # delete default worksheet if it is empty
                if "Sheet1" in worksheet_names:
                    default_worksheet = workbook.get_worksheet_by_name("Sheet1")
                    if not default_worksheet.dim_rowmax:
                        workbook.remove_worksheet(default_worksheet)
                        worksheet_names.remove("Sheet1")

                workbook.close()
                excel_file= send_file(f'{module}.xlsx', as_attachment=True)
                excel_file_path = os.path.abspath(f'{module}.xlsx')
                print(excel_file_path)
                return excel_file

if __name__ == '__main__':
    app.run(debug=True)
